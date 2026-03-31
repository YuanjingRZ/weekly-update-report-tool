import streamlit as st
import pandas as pd
import numpy as np
import io
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


#adding comments to test git branch functionality

st.set_page_config(page_title="Weekly Update Report Tool", page_icon="📊", layout="centered")

st.title("📊 Weekly Update Report Tool")
st.markdown("Upload your three Excel files, set your target numbers, and download the report — no coding needed.")
st.divider()

# ── Step 1: File Uploads ──────────────────────────────────────────────────────
st.subheader("Step 1 · Upload Your Excel Files")
col1, col2, col3 = st.columns(3)
with col1:
    students_file = st.file_uploader("🎒 Students File", type=["xlsx", "xls"], key="students")
with col2:
    adults_file = st.file_uploader("👪 Adults File", type=["xlsx", "xls"], key="adults")
with col3:
    all_file = st.file_uploader("📋 Grant Level File", type=["xlsx", "xls"], key="all")

st.divider()

# ── Step 2: Target values ─────────────────────────────────────────────────────
st.subheader("Step 2 · Set Target Enrollment per Site")
st.info("Enter the target number of students for each site — in the same order as they appear in your Students file. Don't include the Total row.")

if "num_sites" not in st.session_state:
    st.session_state.num_sites = 3

col_add, col_remove = st.columns([1, 1])
with col_add:
    if st.button("➕ Add a site"):
        st.session_state.num_sites += 1
with col_remove:
    if st.button("➖ Remove last site") and st.session_state.num_sites > 1:
        st.session_state.num_sites -= 1

target_values = []
defaults = [152, 200, 100]
for i in range(st.session_state.num_sites):
    default = defaults[i] if i < len(defaults) else 100
    val = st.number_input(f"Site {i+1} — Target # of students", min_value=1, value=default, key=f"target_{i}")
    target_values.append(val)

st.divider()

# ── Step 3: Generate ──────────────────────────────────────────────────────────
st.subheader("Step 3 · Generate Report")

all_uploaded = students_file and adults_file and all_file
if not all_uploaded:
    st.warning("Please upload all three files above before generating.")

if st.button("🚀 Generate Report", disabled=not all_uploaded, type="primary", use_container_width=True):
    with st.status("Running pipeline…", expanded=True) as status:

        try:
            st.write("📂 Reading uploaded files…")
            students_xl  = pd.ExcelFile(students_file)
            adults_bytes = adults_file.read()
            all_bytes    = all_file.read()

            # ── Student Summary Statistics ────────────────────────────────────
            st.write("🍀 Processing Student Summary Statistics…")

            df_part_by_hour = students_xl.parse(1)  # [1] Participants By Hour Band
            df_part_by_hour.columns = df_part_by_hour.iloc[4]
            df_part_by_hour = df_part_by_hour.iloc[5:].reset_index(drop=True)

            daily_site_att = students_xl.parse(0)  # [0] Daily Site Attendance Summary
            daily_site_att.columns = daily_site_att.iloc[2]
            daily_site_att = daily_site_att.iloc[3:]
            daily_site_att.columns.name = None
            daily_site_att = daily_site_att.reset_index(drop=True)
            daily_site_att = daily_site_att[['Total']].iloc[:-1]
            daily_site_att['Total'] = daily_site_att['Total'].str.extract(r'(\d+\.?\d*)')

            all_cols    = ['0', 'Less Than 15', '15-44', '45-89', '90-179', '180-269', '270+']
            served_cols = ['Less Than 15', '15-44', '45-89', '90-179', '180-269', '270+']
            plus15_cols = ['15-44', '45-89', '90-179', '180-269', '270+']
            plus90_cols = ['90-179', '180-269', '270+']
            existing_cols = df_part_by_hour.columns.tolist()
            all_cols    = [c for c in all_cols    if c in existing_cols]
            served_cols = [c for c in served_cols if c in existing_cols]
            plus15_cols = [c for c in plus15_cols if c in existing_cols]
            plus90_cols = [c for c in plus90_cols if c in existing_cols]

            df_part_by_hour[all_cols] = df_part_by_hour[all_cols].apply(pd.to_numeric, errors='coerce').fillna(0)
            df_sub = df_part_by_hour[df_part_by_hour['Site'] == 'Subtotal']

            df_totals = pd.DataFrame({
                '[Total # Enrolled]': df_sub[all_cols].sum(axis=1),
                '[Total # Served]':   df_sub[served_cols].sum(axis=1),
                '[Total # 15+]':      df_sub[plus15_cols].sum(axis=1),
                '[Total # 90+]':      df_sub[plus90_cols].sum(axis=1),
            })
            df_totals.insert(0, '[Target # of students served]', target_values)

            daily_site_att['Total'] = daily_site_att['Total'].astype(int)
            daily_site_att = daily_site_att.rename(columns={'Total': 'Avg. # of Students Per Day'})
            df_totals.insert(3, 'Avg. # of Students Per Day', daily_site_att['Avg. # of Students Per Day'].values)

            school_names = [
                row['Site'] for _, row in df_part_by_hour.iterrows()
                if pd.notna(row['Site']) and row['Site'] != 'Subtotal' and row['Site'] != 'Total'
            ]
            df_totals.insert(0, 'School', school_names)

            total_row = pd.DataFrame(df_totals.iloc[:, 1:].sum()).T
            total_row.insert(0, 'School', 'Total')
            df_totals = pd.concat([df_totals, total_row], ignore_index=True)

            df_totals['# of students 15+ hrs total (% of Target)'] = (
                df_totals['[Total # 15+]'].astype(int).astype(str) + " (" +
                ((df_totals['[Total # 15+]'] / df_totals['[Target # of students served]']) * 100)
                .round().astype(int).astype(str) + "%)"
            )
            df_totals['# of students 90+ hrs total (% of Target)'] = (
                df_totals['[Total # 90+]'].astype(int).astype(str) + " (" +
                ((df_totals['[Total # 90+]'] / df_totals['[Target # of students served]']) * 100)
                .round().astype(int).astype(str) + "%)"
            )
            df_totals = df_totals.drop(columns=['[Total # 15+]', '[Total # 90+]']).reset_index(drop=True)

            # ── Family Component ──────────────────────────────────────────────
            st.write("🌷 Processing Family Component…")
            df_hours = pd.read_excel(io.BytesIO(adults_bytes), sheet_name=2, skiprows=2)  # [2] Participant Attendance Hours
            df_hours['HoursPresent'] = pd.to_numeric(df_hours['HoursPresent'], errors='coerce')
            df_hours['ParticipantId'] = df_hours['ParticipantId'].astype(str).str.replace(r'\.0$', '', regex=True)
            df_active = df_hours[(df_hours['HoursPresent'] > 0) & (df_hours['ParticipantId'].str.len() != 9)]
            result = df_active.groupby('Site')['ParticipantId'].nunique().reset_index()
            result.rename(columns={'ParticipantId': 'Parents Served (Total)'}, inplace=True)
            result.loc[len(result)] = {'Site': 'Total', 'Parents Served (Total)': result['Parents Served (Total)'].sum()}

            # ── Demographics / Missing ────────────────────────────────────────
            st.write("🌸 Processing Participant Demographics…")
            df_part_demo = students_xl.parse(4)  # [4] Participant Demographics
            df_part_demo.columns = df_part_demo.iloc[2]
            df_part_demo = df_part_demo.iloc[3:].reset_index(drop=True)

            def summarize_missing_by_school(df, columns_to_check, category_col='Site'):
                if category_col not in df.columns:
                    raise ValueError(f'{category_col} not found in columns')
                missing_site_rows = df[df[category_col].isna() | (df[category_col].astype(str).str.strip() == '')].copy()
                subset = df[columns_to_check + [category_col]].copy()
                subset_fg = subset[subset[category_col].notna()].copy()
                subset_fg[category_col] = subset_fg[category_col].astype(str).str.title()
                for col in columns_to_check:
                    cleaned = subset_fg[col].astype(str).str.strip()
                    nem = cleaned.str.lower() == 'not entered'
                    if col == 'Gender':
                        vg = cleaned.str.title().isin(['Male', 'Female', 'Non-Binary'])
                        subset_fg[col + '_missing'] = ((~vg) | nem).astype(int)
                    else:
                        subset_fg[col + '_missing'] = (subset_fg[col].isna() | nem).astype(int)
                pid  = df.loc[subset_fg.index, 'ParticipantID'].astype(str).str.strip()
                spid = df.loc[subset_fg.index, 'State ParticipantID'].astype(str).str.strip()
                vp = pid.str.match(r'^[12]\d{8}$')
                vs = spid.str.match(r'^[12]\d{8}$')
                subset_fg['ParticipantID_missing']       = (~vp).astype(int)
                subset_fg['State ParticipantID_missing'] = (~vs).astype(int)
                subset_fg['Both_ID_mismatch']                = (~((pid == spid) & vp & vs)).astype(int)
                missing_cols = (
                    [col + '_missing' for col in columns_to_check]
                    + ['ParticipantID_missing', 'State ParticipantID_missing', 'Both_ID_mismatch']
                )
                pivot = subset_fg.groupby(category_col)[missing_cols].sum().reset_index()
                total_r = pd.DataFrame(pivot[missing_cols].sum()).T
                total_r[category_col] = 'Total'
                pivot = pd.concat([pivot, total_r], ignore_index=True)
                all_mf = df.copy()
                pid_a  = all_mf['ParticipantID'].astype(str).str.strip()
                spid_a = all_mf['State ParticipantID'].astype(str).str.strip()
                vpa = pid_a.str.match(r'^[12]\d{8}$')
                vsa = spid_a.str.match(r'^[12]\d{8}$')
                for col in columns_to_check:
                    cleaned = all_mf[col].astype(str).str.strip()
                    nem = cleaned.str.lower() == 'not entered'
                    if col == 'Gender':
                        vg = cleaned.str.title().isin(['Male', 'Female', 'Non-Binary'])
                        all_mf[col + '_missing'] = ((~vg) | nem).astype(int)
                    else:
                        all_mf[col + '_missing'] = (all_mf[col].isna() | nem).astype(int)
                all_mf['ParticipantID_missing']       = (~vpa).astype(int)
                all_mf['State ParticipantID_missing'] = (~vsa).astype(int)
                all_mf['Both_ID_mismatch']                = (~((pid_a == spid_a) & vpa & vsa)).astype(int)
                dob_parsed = pd.to_datetime(all_mf['Date Of Birth'], errors='coerce')
                all_mf['DOB_too_young'] = (dob_parsed.dt.year > 2023).astype(int)
                flag_cols2 = [col + '_missing' for col in columns_to_check] + ['ParticipantID_missing']
                total_missing_rows = all_mf[all_mf[flag_cols2].sum(axis=1) > 0].copy()
                young_dob_rows = all_mf[all_mf['DOB_too_young'] == 1].copy()
                return pivot, missing_site_rows, total_missing_rows, flag_cols2, young_dob_rows

            columns_of_interest = ['Date Of Birth', 'Grade Level', 'Gender']
            missing_summary, missing_site_rows, total_missing_rows, flag_cols, young_dob_rows = \
                summarize_missing_by_school(df_part_demo, columns_of_interest)

            # ── Site Summary Report ───────────────────────────────────────────
            st.write("🪻 Building Site Summary Report…")
            all_io = io.BytesIO(all_bytes)
            df_act = pd.read_excel(all_io, sheet_name=3, skiprows=2)  # [3] Activity-Session Details
            all_io.seek(0)
            df_enr = pd.read_excel(all_io, sheet_name=5, skiprows=2)  # [5] Session Enrollment by Session
            all_io.seek(0)
            df_att = pd.read_excel(all_io, sheet_name=6, skiprows=4)  # [6] Daily Activity Attendance Summa

            cols_act = ['Site', 'Activity', 'Session', 'Days Scheduled', 'Session Start Date']
            df_act = df_act[cols_act].copy()
            df_act['Days Scheduled'] = pd.to_numeric(df_act['Days Scheduled'], errors='coerce')
            df_act['Session Start Date'] = pd.to_datetime(df_act['Session Start Date'], errors='coerce')
            today = pd.Timestamp.today().normalize()

            df_enr = df_enr[['Site', 'Activity', 'Session', 'Enrolled Count']].copy()
            df_enr['Enrolled Count'] = pd.to_numeric(df_enr['Enrolled Count'], errors='coerce')
            df_enr.rename(columns={'Enrolled Count': 'Enrolled Participant'}, inplace=True)

            def extract_average(val):
                if pd.isna(val): return np.nan
                try: return float(str(val).replace('Average:', '').strip())
                except: return np.nan

            df_att = df_att[['Site', 'Activity', 'Session', 'Total']].copy()
            df_att['Total'] = df_att['Total'].apply(extract_average)
            df_att.rename(columns={'Total': 'Average Daily Attendance'}, inplace=True)

            sites = [
                s for s in df_act['Site'].dropna().unique()
                if str(s).strip() != '' and not str(s).startswith('Total') and not str(s).startswith('Average')
            ]
            site_tables = {}
            for site in sites:
                m = pd.merge(df_act[df_act['Site'] == site], df_enr[df_enr['Site'] == site],
                             on=['Site', 'Activity', 'Session'], how='outer')
                m = pd.merge(m, df_att[df_att['Site'] == site],
                             on=['Site', 'Activity', 'Session'], how='outer')
                m = m[~(m['Session Start Date'] >= today)].drop(columns=['Session Start Date'], errors='ignore')
                m = m.fillna('-').sort_values(['Activity', 'Session']).reset_index(drop=True)
                site_tables[site] = m

            # ── Write Excel ───────────────────────────────────────────────────
            st.write("🌈 Writing Excel report…")
            output_buffer = io.BytesIO()
            with pd.ExcelWriter(output_buffer, engine='openpyxl') as writer:
                df_totals.to_excel(writer, sheet_name='Student Statistics', index=False)
                result.to_excel(writer, sheet_name='Parents Served Summary', index=False)
                missing_summary.to_excel(writer, sheet_name='Missing Summary', index=False)
                missing_site_rows.to_excel(writer, sheet_name='Missing Site Info', index=False)
                display_cols = [c for c in total_missing_rows.columns if not c.endswith('_missing') and c != 'DOB_too_young']
                total_missing_rows[display_cols].to_excel(writer, sheet_name='Total Missing Info', index=False)
                young_dc = [c for c in young_dob_rows.columns if not c.endswith('_missing') and c != 'DOB_too_young']
                young_dob_rows[young_dc].to_excel(writer, sheet_name='Young DOB', index=False)
                for site_name, final_df in site_tables.items():
                    safe = str(site_name)[:31].replace(':', '').replace('/', '').replace('\\', '').replace('?', '').replace('*', '')
                    final_df.to_excel(writer, sheet_name=safe, index=False)

            # Apply highlights
            output_buffer.seek(0)
            wb = load_workbook(output_buffer)
            red_fill  = PatternFill('solid', start_color='FF9999', end_color='FF9999')
            blue_fill = PatternFill('solid', start_color='9999FF', end_color='9999FF')
            flag_to_original = {fc: fc[:-len('_missing')] for fc in flag_cols if fc.endswith('_missing')}
            ws = wb['Total Missing Info']
            header = {cell.value: cell.column for cell in ws[1]}
            for row_idx, (_, row) in enumerate(total_missing_rows.iterrows(), start=2):
                for flag_col, orig_col in flag_to_original.items():
                    if orig_col in header and row.get(flag_col, 0) == 1:
                        ws.cell(row=row_idx, column=header[orig_col]).fill = red_fill
            ws2 = wb['Young DOB']
            header2 = {cell.value: cell.column for cell in ws2[1]}
            if 'Date Of Birth' in header2:
                dob_col_idx = header2['Date Of Birth']
                for row_idx in range(2, len(young_dob_rows) + 2):
                    ws2.cell(row=row_idx, column=dob_col_idx).fill = blue_fill

            final_buffer = io.BytesIO()
            wb.save(final_buffer)
            final_buffer.seek(0)

            status.update(label="✅ Report ready!", state="complete")
            st.session_state.output_bytes = final_buffer.read()
            st.session_state.report_ready = True

        except Exception as e:
            status.update(label="❌ Something went wrong", state="error")
            st.error(f"Error: {e}")
            st.session_state.report_ready = False

if st.session_state.get("report_ready"):
    st.success("🎉 Your report is ready!")
    st.download_button(
        label="⬇️ Download Weekly_Update_Report.xlsx",
        data=st.session_state.output_bytes,
        file_name="Weekly_Update_Report.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        type="primary",
    )

    st.divider()

    # ── Step 4: Highlight Participant IDs ─────────────────────────────────────
    st.subheader("Step 4 · Highlight Participant IDs (Optional)")
    st.info(
        "Want to pull out and highlight **ParticipantID** or **State ParticipantID** columns "
        "in the *Total Missing Info* sheet? Pick your options below and download a highlighted version."
    )

    col_pid, col_spid = st.columns(2)
    with col_pid:
        highlight_pid = st.checkbox("Highlight **ParticipantID**", value=False)
        pid_color = st.color_picker("ParticipantID highlight color", value="#FFFF00", disabled=not highlight_pid)
    with col_spid:
        highlight_spid = st.checkbox("Highlight **State ParticipantID**", value=False)
        spid_color = st.color_picker("State ParticipantID highlight color", value="#90EE90", disabled=not highlight_spid)

    if highlight_pid or highlight_spid:
        if st.button("✨ Apply Highlights & Download", type="primary", use_container_width=True):
            try:
                wb2 = load_workbook(io.BytesIO(st.session_state.output_bytes))
                ws_tmi = wb2["Total Missing Info"]
                header_tmi = {cell.value: cell.column for cell in ws_tmi[1]}
                max_row = ws_tmi.max_row

                def hex_to_openpyxl(hex_color):
                    """Convert #RRGGBB to openpyxl RRGGBB (no #)."""
                    return hex_color.lstrip("#").upper()

                if highlight_pid and "ParticipantID" in header_tmi:
                    fill_pid = PatternFill("solid",
                                          start_color=hex_to_openpyxl(pid_color),
                                          end_color=hex_to_openpyxl(pid_color))
                    col_idx = header_tmi["ParticipantID"]
                    for r in range(2, max_row + 1):
                        ws_tmi.cell(row=r, column=col_idx).fill = fill_pid

                if highlight_spid and "State ParticipantID" in header_tmi:
                    fill_spid = PatternFill("solid",
                                           start_color=hex_to_openpyxl(spid_color),
                                           end_color=hex_to_openpyxl(spid_color))
                    col_idx = header_tmi["State ParticipantID"]
                    for r in range(2, max_row + 1):
                        ws_tmi.cell(row=r, column=col_idx).fill = fill_spid

                highlighted_buffer = io.BytesIO()
                wb2.save(highlighted_buffer)
                highlighted_buffer.seek(0)

                st.download_button(
                    label="⬇️ Download Highlighted Report",
                    data=highlighted_buffer.read(),
                    file_name="Weekly_Update_Report_Highlighted.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            except Exception as e:
                st.error(f"Could not apply highlights: {e}")
    else:
        st.caption("☝️ Check at least one option above to enable the highlight download.")